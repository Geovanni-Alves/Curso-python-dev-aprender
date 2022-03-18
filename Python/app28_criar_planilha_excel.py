import openpyxl
planilha_test = openpyxl.Workbook()
planilha_test.create_sheet("Frutas")
planilha_test.create_sheet("Legumes")
planilha_test.create_sheet("Sementes")
# print(planilha_test.sheetnames)

cabecalho_frutas = ["Titulo", "Localização", "Preço"]

sheet_frutas = planilha_test["Frutas"]
sheet_legumes = planilha_test["Legumes"]
sheet_sementes = planilha_test["Sementes"]

sheet_frutas.append(cabecalho_frutas)

frutas = [["Uva", "Mercado", 5],
          ["Melancia", "Mercado", 15], ["Bolo", "Mercado", 40]]

for fruta in frutas:
    sheet_frutas.append(fruta)


planilha_test.save("Dados Supermercado.xlsx")
