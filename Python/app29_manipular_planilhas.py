import openpyxl

planilha = openpyxl.load_workbook("pessoas.xlsx")
print(planilha.sheetnames)
sheet1 = planilha["Sheet1"]
print(sheet1["C3"].value)
sheet1["C3"].value = "Hashirama"
print(sheet1["C3"].value)
# planilha.save("pessoasv2.xlsx")

sheet1.cell(row=3, column=3, value="Mark")
print(sheet1["C3"].value)
for linha in sheet1.iter_rows(min_row=2, max_row=10, min_col=2, max_col=4):
    print(linha[0].value, linha[1].value, linha[2].value)

for linha in sheet1.iter_cols(min_col=2, max_col=2, min_row=2):
    for cell in linha:
        print(cell.value)
