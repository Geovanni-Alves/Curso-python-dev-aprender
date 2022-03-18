import openpyxl

planilha = openpyxl.load_workbook("pessoas.xlsx")
sheet1 = planilha["Sheet1"]
print(sheet1["B4"].value)
sheet1["B4"].value = "Jimmy"
print(sheet1["B4"].value)

for linha in sheet1.iter_rows(min_row=2, max_row=10):
    print(linha[0].value, linha[1].value)

for linha in sheet1.iter_cols(min_row=2, min_col=3, max_col=3):
    for cell in linha:
        print(cell.value)
